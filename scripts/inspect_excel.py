import openpyxl
from openpyxl.utils import get_column_letter

path = "docs/3-H.소프트웨어 버전 및 배포관리(최종)_26.04.14.xlsx"
wb = openpyxl.load_workbook(path, data_only=True)

print("=== Sheets ===")
for ws in wb.worksheets:
    print(f"- {ws.title!r}  rows={ws.max_row}  cols={ws.max_column}")

target = None
for ws in wb.worksheets:
    if "SW" in ws.title or "배포" in ws.title:
        target = ws
        break
if target is None:
    target = wb.worksheets[0]

print(f"\n=== Sheet: {target.title!r} ===")
for r in range(1, min(target.max_row, 80) + 1):
    row_vals = []
    for c in range(1, target.max_column + 1):
        v = target.cell(row=r, column=c).value
        if v is None:
            row_vals.append("")
        else:
            row_vals.append(str(v))
    if any(v.strip() for v in row_vals):
        print(f"R{r}: {row_vals}")

print("\n=== Merged ranges ===")
for mr in list(target.merged_cells.ranges)[:30]:
    print(mr)
